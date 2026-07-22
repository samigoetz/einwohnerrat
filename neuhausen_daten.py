#!/usr/bin/env python3
"""
Holt die Vorstoss- und Berichtsseiten der Gemeinde Neuhausen am Rheinfall
und schreibt die aufbereiteten Daten als neuhausen_daten.js in denselben
Ordner. Die Datei neuhausen-monitor.html liest diese Daten automatisch.

Abhaengigkeiten:  pip install requests beautifulsoup4
Aufruf:           python3 neuhausen_daten.py
Automatisierung:  per Cron oder launchd regelmaessig ausfuehren.
"""

import json
import re
import shutil
import subprocess
import sys
import tempfile
import time
import unicodedata
from datetime import datetime
from pathlib import Path

try:
    from zoneinfo import ZoneInfo
    _ZEITZONE = ZoneInfo("Europe/Zurich")
except Exception:
    _ZEITZONE = None

import requests
from bs4 import BeautifulSoup

BASIS = Path(__file__).resolve().parent
AUSGABE = BASIS / "neuhausen_daten.js"

VORSTOSS_QUELLEN = [
    {"art": "Kleine Anfrage", "kuerzel": "ka", "url": "https://neuhausen.ch/kleine_anfrage"},
    {"art": "Interpellation", "kuerzel": "ip", "url": "https://neuhausen.ch/interpellationen"},
    {"art": "Postulat",       "kuerzel": "po", "url": "https://neuhausen.ch/postulate"},
    {"art": "Motion",         "kuerzel": "mo", "url": "https://neuhausen.ch/motionen"},
]
BA_URL = "https://neuhausen.ch/berichte_antraege"
BP_URL = "https://neuhausen.ch/beschluesse_protokolle"
AKTUELLES_URL = "https://neuhausen.ch/aktuelles"
PRESSE_SHN_URL = "https://www.shn.ch/region/neuhausen"
PRESSE_SHAZ_FEED = "https://www.shaz.ch/feed/"
PRESSE_GOOGLE_FEED = ("https://news.google.com/rss/search?"
                      "q=%22Neuhausen+am+Rheinfall%22&hl=de-CH&gl=CH&ceid=CH:de")
PRESSE_ARCHIV_AUSGABE = BASIS / "presse_archiv.js"
PRESSE_ANZEIGE_TAGE = 365        # juengere Artikel stehen direkt auf der Seite
PRESSE_LEAD_MAX_PRO_LAUF = 40    # so viele fehlende Leads werden pro Lauf geholt
PRESSE_RUECKFUELL_AB_JAHR = 2020
KENNZAHLEN_AUSGABE = BASIS / "kennzahlen.js"
KENNZAHLEN_PRUEFTAKT_TAGE = 7   # amtliche Zahlen aendern sich selten
KENNZAHLEN_VERSION = 14          # bei Ausbau/Korrektur erhoehen: erzwingt Neuabfrage
STATTAB_BASIS = "https://www.pxweb.bfs.admin.ch/api/v1/de"
STATTAB_SEITE = "https://www.pxweb.bfs.admin.ch/pxweb/de"
FEED_AUSGABE = BASIS / "feed.xml"
VOLLTEXT_AUSGABE = BASIS / "volltext.js"
VOLLTEXT_MAX_ZEICHEN = 150000   # Textobergrenze pro Dokument
OCR_MAX_SEITEN = 40             # OCR hoechstens fuer so viele Seiten pro PDF
OCR_AUFLOESUNG = 200            # dpi fuer die Texterkennung

# ---------------------------------------------------------------------------
# FRAKTIONSZUORDNUNG  (bitte selbst ergaenzen und korrigieren!)
#
# Format:  "Vorname Nachname": "Fraktion",
# Quelle aktuelle Mitglieder: https://neuhausen.ch/einwohnerrat_mitglieder
# Historische Zuordnungen stammen aus Parteiangaben in den Vorstosstiteln.
# Personen ohne Eintrag erscheinen als "Unbekannt"; das Skript listet sie
# nach jedem Lauf im Terminal auf.
# Hinweis: Gemaess Sitzverteilung bilden SP und parteilos eine Fraktion.
# Wer das zusammenfassen will, ersetzt "parteilos" unten durch "SP".
# ---------------------------------------------------------------------------
FRAKTIONEN = {
    # --- Aktuelle Mitglieder (Legislatur 2025-2028) ---
    "Roland Müller": "Grüne",
    "Nina Schärrer": "FDP",
    "Fabian Bolli": "GLP",
    "Urim Dakaj": "SP",
    "Oliver Fessler": "SVP",
    "Peter Fischli": "FDP",
    "Herbert Hirsiger": "SVP",
    "Arnold Isliker": "SVP",
    "Deborah Isliker": "SVP",
    "Melanie Knuchel": "SP",
    "Matthias Koch": "GLP",
    "Bernhard Koller": "EDU",
    "Thomas Leuzinger": "SP",
    "Dimitrij Ruh": "SP",
    "Christian Schenk": "SP",
    "Ernst Schläpfer": "SP",
    "Silvia Schlegel": "Die Mitte",
    "Urs Schüpbach": "parteilos",
    "Marco Torsello": "FDP",
    "Isabella Zellweger": "SVP",
    # --- Fruehere Mitglieder (Partei aus Vorstosstiteln belegt) ---
    "Jakob Walter": "SP",
    "Renzo Loiudice": "SP",
    "August Hafner": "SP",
    "Priska Weber-Widmer": "SP",
    "Daniel Borer": "SP",
    "Willi Josel": "SVP",
    "Peter Schmid": "SVP",
    "Walter Herrmann": "FDP",
    "Walter Hermann": "FDP",  # Schreibvariante auf der Gemeindeseite
    "Felix Tenger": "FDP",
    "Thomas Theiler": "Die Mitte",   # damals CVP
    "Marcel Stettler": "Die Mitte",  # damals CVP
    "Rita Flück Hänzi": "Die Mitte", # damals CVP
    "Urs Hinnen": "ÖBS",
    # --- Ergaenzungen gemaess Angaben SP Neuhausen ---
    "Nicole Hinder": "AL",
    "Adrian Schüpbach": "SVP",
    "Randy Ruh": "GLP",
    "Sabina Tektas-Sorg": "SP",
    "Sara Jucker": "SVP",
    "Andreas Neuenschwander": "SVP",
    "Ruedi Meier": "SP",
    "Rolf Forster": "SVP",
    "Robert Eichmann": "SVP",
    "Markus Anderegg": "FDP",
    "Michael Bernath": "ÖBS",
    "René Sauzet": "FDP",
    # --- Schreibvarianten auf der Gemeindeseite ---
    "Bernahrd Koller": "EDU",   # Tippfehler fuer Bernhard Koller
    "Sarah Jucker": "SVP",      # Variante von Sara Jucker
    # --- Keine Personen: eigene Kategorien in der Statistik ---
    "GPK": "GPK",               # Geschaeftspruefungskommission
    "Volksmotion": "Volksmotion",
}

_NAME_TITEL_RE = re.compile(r"^(dr|prof|med)\.?\s+", re.I)

# Personen, die bei einem Lauf nicht zugeordnet werden konnten
UNBEKANNTE_PERSONEN = set()


def _norm_name(name: str) -> str:
    """Titel wie 'Dr.' entfernen, Leerraum normalisieren."""
    n = sauber(name)
    n = _NAME_TITEL_RE.sub("", n)
    return n


def fraktionen_fuer(person_feld: str) -> list:
    """
    Ermittelt die Fraktion(en) fuer das Personenfeld eines Vorstosses.
    Mehrere Einreichende ("A und B", "A / B") werden einzeln zugeordnet.
    Nicht zuordenbare Namen landen in UNBEKANNTE_PERSONEN.
    """
    if not person_feld or person_feld == "-":
        return ["Unbekannt"]
    teile = re.split(r"\s+und\s+|\s*/\s*|\s*&\s*|\s*,\s*", person_feld)
    gefunden = []
    for teil in teile:
        n = _norm_name(teil)
        if not n:
            continue
        # Sammelbegriffe ohne Fraktionsaussage ueberspringen
        if re.match(r"^mitunterzeichn|^weitere$|^andere$", n, re.I):
            continue
        frak = FRAKTIONEN.get(n)
        if frak is None and re.match(r"^[A-Za-zÀ-ÿ]\.\s*\S", n):
            # Nur echte Kurzformen wie "R. Müller" ueber den Nachnamen aufloesen
            for voll, f in FRAKTIONEN.items():
                nachname = voll.split()[-1]
                if n.endswith(" " + nachname) or n.split()[-1] == nachname:
                    frak = f
                    break
        if frak is None:
            UNBEKANNTE_PERSONEN.add(n)
        gefunden.append(frak or "Unbekannt")
    einzig = []
    for f in gefunden:
        if f not in einzig:
            einzig.append(f)
    return einzig or ["Unbekannt"]

HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                   "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Safari/605.1.15"),
    "Accept-Language": "de-CH,de;q=0.9",
}
TIMEOUT = 30
DATUM_RE = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")


def absolut(href: str) -> str:
    if href.startswith("http"):
        return href
    if href.startswith("/"):
        return "https://neuhausen.ch" + href
    return "https://neuhausen.ch/" + href


def sauber(text: str) -> str:
    text = re.sub(r"\s+", " ", text or "").strip()
    # Zerlegte Umlaute (a + Trema) zu normalen Umlauten zusammenfassen
    return unicodedata.normalize("NFC", text)


def hole(url: str) -> str:
    resp = requests.get(url, headers=HEADERS, timeout=TIMEOUT)
    resp.raise_for_status()
    # Zeichensatz: Header-Angabe respektieren, sonst UTF-8 annehmen.
    ctype = resp.headers.get("content-type", "")
    if "charset" not in ctype.lower():
        resp.encoding = "utf-8"
    return resp.text


def _zelle(zeile, klasse):
    """Text einer Pseudo-Zelle (span.td.row_X) innerhalb einer Zeile."""
    el = zeile.find("span", class_=lambda c: c and "td" in c.split() and klasse in c.split())
    return sauber(el.get_text()) if el else ""


def _zellen_element(zeile, klasse):
    return zeile.find("span", class_=lambda c: c and "td" in c.split() and klasse in c.split())


def parse_vorstoesse(html: str, quelle: dict) -> list:
    """
    Die Seiten bilden Tabellen mit div/span nach:
      div.tr  >  span.td.row_1 (Datum), row_3 (Name), row_2 (Nummer),
                 row_4 (Inhalt mit PDF-Links), row_5 (Status),
                 row_6 (Duplikat fuer Mobilansicht, wird ignoriert)
    """
    soup = BeautifulSoup(html, "html.parser")
    ergebnis = []
    gesehen = set()

    for zeile in soup.find_all("div", class_="tr"):
        datum = _zelle(zeile, "row_1")
        if not DATUM_RE.match(datum):
            continue  # Kopfzeilen (span.th) und Fremdes ueberspringen

        person = _zelle(zeile, "row_3")
        nummer = _zelle(zeile, "row_2")
        status = _zelle(zeile, "row_5")

        inhalt = _zellen_element(zeile, "row_4")
        haupt, antwort = None, None
        if inhalt is not None:
            for a in inhalt.find_all("a", href=True):
                if "/fileupload/" not in a["href"]:
                    continue
                t = sauber(a.get_text())
                u = absolut(a["href"])
                if haupt is None:
                    haupt = {"titel": t, "url": u}
                elif antwort is None and re.search(r"beantwortung|antwort", t, re.I):
                    antwort = {"titel": t, "url": u}

        titel = haupt["titel"] if haupt else (sauber(inhalt.get_text()) if inhalt else "")
        if not titel:
            continue
        url = haupt["url"] if haupt else quelle["url"]

        schluessel = f"{quelle['kuerzel']}|{datum}|{nummer}|{titel}"
        if schluessel in gesehen:
            continue
        gesehen.add(schluessel)

        ergebnis.append({
            "schluessel": schluessel,
            "datum": datum,
            "person": person or "-",
            "nummer": nummer,
            "art": quelle["art"],
            "kuerzel": quelle["kuerzel"],
            "titel": titel,
            "url": url,
            "antwortUrl": antwort["url"] if antwort else None,
            "status": status,
            "fraktionen": fraktionen_fuer(person),
        })
    return ergebnis


def parse_berichte(html: str) -> list:
    """
    Gleiche Pseudo-Tabellenstruktur (div.tr > span.td.row_X), aber andere
    Spalten: Datum, Art, Inhalt, Status (keine Person). Da die row-Nummern
    abweichen koennen, wird flexibel gearbeitet:
      Datum  = Zelle mit Datumsformat
      Inhalt = Zelle mit den meisten fileupload-Links
      Status = letzte kurze Textzelle (pendent/erledigt/...)
      Art    = uebrige Textzelle
    Duplikat-Zellen (Mobilansicht) werden ueber den Schluessel entschaerft.
    """
    soup = BeautifulSoup(html, "html.parser")
    ergebnis = []
    gesehen = set()
    STATUS_WORTE = re.compile(r"pendent|erledigt|zur(ü|ue)ck", re.I)

    for zeile in soup.find_all("div", class_="tr"):
        zellen = zeile.find_all(
            "span", class_=lambda c: c and "td" in c.split()
        )
        if len(zellen) < 3:
            continue

        datum, art, status = "", "", ""
        inhalt_zelle, max_links = None, 0

        for td in zellen:
            text = sauber(td.get_text())
            n_links = len([a for a in td.find_all("a", href=True)
                           if "/fileupload/" in a["href"]])
            if n_links > max_links:
                max_links, inhalt_zelle = n_links, td
                continue
            if not datum and DATUM_RE.match(text):
                datum = text
            elif STATUS_WORTE.search(text) and len(text) < 40:
                status = text
            elif not art and text and n_links == 0:
                art = text

        if not datum or inhalt_zelle is None:
            continue

        dokumente, seen2 = [], set()
        for a in inhalt_zelle.find_all("a", href=True):
            if "/fileupload/" not in a["href"]:
                continue
            t = sauber(a.get_text())
            u = absolut(a["href"])
            if (t, u) in seen2:
                continue
            seen2.add((t, u))
            dokumente.append({"titel": t, "url": u})
        if not dokumente:
            continue

        schluessel = f"ba|{datum}|{dokumente[0]['titel']}"
        if schluessel in gesehen:
            continue
        gesehen.add(schluessel)

        ergebnis.append({
            "schluessel": schluessel,
            "datum": datum,
            "art": art or "Bericht und Antrag",
            "status": status,
            "haupt": dokumente[0],
            "beilagen": dokumente[1:],
        })
    return ergebnis


def parse_beschluesse(html: str) -> list:
    """
    Beschluesse & Protokolle: gleiche Pseudo-Tabellenstruktur, Spalten
    Datum, Nummer, Inhalt. Da die row-Nummern abweichen koennen, wird
    flexibel gearbeitet:
      Datum  = Zelle mit Datumsformat
      Inhalt = Zelle mit den meisten fileupload-Links
      Nummer = kuerzeste uebrige Textzelle ohne Links
    """
    soup = BeautifulSoup(html, "html.parser")
    ergebnis = []
    gesehen = set()

    for zeile in soup.find_all("div", class_="tr"):
        zellen = zeile.find_all(
            "span", class_=lambda c: c and "td" in c.split()
        )
        if len(zellen) < 2:
            continue

        datum, nummer = "", ""
        inhalt_zelle, max_links = None, 0
        text_zellen = []

        for td in zellen:
            text = sauber(td.get_text())
            n_links = len([a for a in td.find_all("a", href=True)
                           if "/fileupload/" in a["href"]])
            if n_links > max_links:
                max_links, inhalt_zelle = n_links, td
                continue
            if not datum and DATUM_RE.match(text):
                datum = text
            elif text and n_links == 0:
                text_zellen.append(text)

        if not datum:
            continue
        if text_zellen:
            nummer = min(text_zellen, key=len)

        dokumente, seen2 = [], set()
        if inhalt_zelle is not None:
            for a in inhalt_zelle.find_all("a", href=True):
                if "/fileupload/" not in a["href"]:
                    continue
                t = sauber(a.get_text())
                u = absolut(a["href"])
                if (t, u) in seen2:
                    continue
                seen2.add((t, u))
                dokumente.append({"titel": t, "url": u})

        if dokumente:
            haupt = dokumente[0]
            beilagen = dokumente[1:]
        else:
            # Zeile ohne Dokumentlinks: laengste Textzelle als Titel verwenden
            lange = [t for t in text_zellen if t != nummer]
            if not lange:
                continue
            haupt = {"titel": max(lange, key=len), "url": BP_URL}
            beilagen = []

        schluessel = f"bp|{datum}|{nummer}|{haupt['titel']}"
        if schluessel in gesehen:
            continue
        gesehen.add(schluessel)

        ergebnis.append({
            "schluessel": schluessel,
            "datum": datum,
            "nummer": nummer,
            "haupt": haupt,
            "beilagen": beilagen,
        })
    return ergebnis


# ---------------------------------------------------------------------------
# Volltext-Erfassung: extrahiert den Text aller verlinkten PDFs fuer die
# Volltextsuche. Bereits verarbeitete Dokumente werden aus volltext.js
# wiederverwendet (Zwischenspeicher), nur Neues wird heruntergeladen.
# Scans ohne Textebene werden per Texterkennung (tesseract) erfasst,
# sofern die Werkzeuge vorhanden sind (auf GitHub: ja).
# ---------------------------------------------------------------------------

def _lade_volltext_cache() -> dict:
    if not VOLLTEXT_AUSGABE.exists():
        return {}
    try:
        roh = VOLLTEXT_AUSGABE.read_text(encoding="utf-8")
        start = roh.find("{")
        ende = roh.rfind("}")
        cache = json.loads(roh[start:ende + 1])
        return cache if isinstance(cache, dict) else {}
    except Exception:
        return {}


def _normalisiere_volltext(text: str) -> str:
    text = unicodedata.normalize("NFC", text or "")
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text[:VOLLTEXT_MAX_ZEICHEN]


def _pdf_text(daten: bytes) -> str:
    """Textebene eines PDFs auslesen (leer, wenn keine vorhanden)."""
    from pypdf import PdfReader
    import io
    leser = PdfReader(io.BytesIO(daten))
    if getattr(leser, "is_encrypted", False):
        try:
            leser.decrypt("")
        except Exception:
            return ""
    teile = []
    for seite in leser.pages:
        teile.append(seite.extract_text() or "")
        if sum(len(t) for t in teile) > VOLLTEXT_MAX_ZEICHEN:
            break
    return " ".join(teile)


def _ocr_moeglich() -> bool:
    return bool(shutil.which("tesseract") and shutil.which("pdftoppm"))


def _ocr_text(daten: bytes) -> str:
    """Texterkennung fuer Scans: PDF zu Bildern, dann tesseract (Deutsch)."""
    with tempfile.TemporaryDirectory() as ordner:
        pfad = Path(ordner)
        (pfad / "dok.pdf").write_bytes(daten)
        subprocess.run(
            ["pdftoppm", "-r", str(OCR_AUFLOESUNG), "-png",
             "-l", str(OCR_MAX_SEITEN), "dok.pdf", "seite"],
            cwd=ordner, check=True, capture_output=True, timeout=300,
        )
        teile = []
        for bild in sorted(pfad.glob("seite*.png")):
            ergebnis = subprocess.run(
                ["tesseract", str(bild), "stdout", "-l", "deu"],
                capture_output=True, timeout=300,
            )
            teile.append(ergebnis.stdout.decode("utf-8", errors="ignore"))
            if sum(len(t) for t in teile) > VOLLTEXT_MAX_ZEICHEN:
                break
        return " ".join(teile)


def _hole_pdf(url: str) -> bytes:
    resp = requests.get(url, headers=HEADERS, timeout=60)
    resp.raise_for_status()
    return resp.content


def _sammle_dokument_urls(vorstoesse, berichte, beschluesse) -> list:
    urls = []
    gesehen = set()

    def nimm(u):
        if u and "/fileupload/" in u and u not in gesehen:
            gesehen.add(u)
            urls.append(u)

    for v in vorstoesse:
        nimm(v.get("url"))
        nimm(v.get("antwortUrl"))
    for gruppe in (berichte, beschluesse):
        for e in gruppe:
            nimm(e.get("haupt", {}).get("url"))
            for b in e.get("beilagen", []):
                nimm(b.get("url"))
    return urls


def baue_volltext(vorstoesse, berichte, beschluesse) -> None:
    try:
        import pypdf  # noqa: F401
    except ImportError:
        print("Volltext uebersprungen: pypdf fehlt "
              "(Installation: python3 -m pip install --user pypdf)")
        return

    cache = _lade_volltext_cache()
    urls = _sammle_dokument_urls(vorstoesse, berichte, beschluesse)
    kann_ocr = _ocr_moeglich()

    # Erneut versuchen, was frueher scheiterte oder mangels OCR offen blieb
    offen = [u for u in urls
             if u not in cache
             or cache[u].get("q") in ("fehler",)
             or (cache[u].get("q") == "ocr_fehlt" and kann_ocr)]

    neu = ocr_n = leer = fehler = 0
    for i, url in enumerate(offen, 1):
        try:
            daten = _hole_pdf(url)
            text = _pdf_text(daten)
            if len(text.strip()) >= 50:
                cache[url] = {"t": _normalisiere_volltext(text), "q": "pdf"}
            elif kann_ocr:
                try:
                    ocr = _ocr_text(daten)
                    if len(ocr.strip()) >= 50:
                        cache[url] = {"t": _normalisiere_volltext(ocr), "q": "ocr"}
                        ocr_n += 1
                    else:
                        cache[url] = {"t": "", "q": "leer"}
                        leer += 1
                except Exception:
                    cache[url] = {"t": "", "q": "fehler"}
                    fehler += 1
            else:
                cache[url] = {"t": "", "q": "ocr_fehlt"}
            neu += 1
        except Exception:
            cache[url] = {"t": "", "q": "fehler"}
            fehler += 1
        if i % 25 == 0 or i == len(offen):
            print(f"  Volltext: {i}/{len(offen)} neue Dokumente verarbeitet")
        time.sleep(0.15)

    js = "window.NEUHAUSEN_VOLLTEXT = " + json.dumps(cache, ensure_ascii=False) + ";\n"
    VOLLTEXT_AUSGABE.write_text(js, encoding="utf-8")

    erfasst = sum(1 for e in cache.values() if e.get("q") in ("pdf", "ocr"))
    ohne = sum(1 for e in cache.values() if e.get("q") == "leer")
    ausstehend = sum(1 for e in cache.values() if e.get("q") == "ocr_fehlt")
    print(f"Volltext: {erfasst} von {len(cache)} Dokumenten erfasst "
          f"(neu: {neu}, davon OCR: {ocr_n}, ohne Text: {ohne}, Fehler: {fehler})")
    if ausstehend and not kann_ocr:
        print(f"  Hinweis: {ausstehend} Scans warten auf Texterkennung; "
              "die passiert automatisch beim naechsten GitHub-Lauf "
              "(lokal: tesseract und poppler installieren).")


_WOCHENTAGE = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
_MONATE = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
           "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _rfc822(datum: str) -> str:
    """dd.mm.yyyy -> RFC-822-Datum (12:00 Schweizer Zeit), fuer RSS."""
    m = re.match(r"^(\d{2})\.(\d{2})\.(\d{4})$", datum)
    if not m:
        return ""
    dt = datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)), 12, 0,
                  tzinfo=_ZEITZONE)
    offset = dt.strftime("%z") or "+0000"
    return (f"{_WOCHENTAGE[dt.weekday()]}, {dt.day:02d} "
            f"{_MONATE[dt.month - 1]} {dt.year} 12:00:00 {offset}")


def baue_feed(vorstoesse: list, berichte: list, beschluesse: list, aktuelles: list, presse: list) -> str:
    """Erzeugt einen RSS-2.0-Feed mit den neuesten Eintraegen aller Bereiche."""
    from xml.sax.saxutils import escape

    def datum_zahl(d):
        m = re.match(r"^(\d{2})\.(\d{2})\.(\d{4})$", d)
        return int(m.group(3) + m.group(2) + m.group(1)) if m else 0

    eintraege = []
    for v in vorstoesse:
        eintraege.append({
            "titel": f"[{v['art']}] {v['titel']}",
            "url": v["url"], "datum": v["datum"], "id": v["schluessel"],
        })
    for b in berichte:
        eintraege.append({
            "titel": f"[Bericht & Antrag] {b['haupt']['titel']}",
            "url": b["haupt"]["url"], "datum": b["datum"], "id": b["schluessel"],
        })
    for p in beschluesse:
        eintraege.append({
            "titel": f"[Beschluss/Protokoll] {p['haupt']['titel']}",
            "url": p["haupt"]["url"], "datum": p["datum"], "id": p["schluessel"],
        })
    for a in aktuelles:
        eintraege.append({
            "titel": f"[Aktuelles] {a['titel']}",
            "url": AKTUELLES_URL, "datum": a["datum"], "id": a["schluessel"],
        })
    for p in presse:
        eintraege.append({
            "titel": f"[Presse \u00b7 {p['quelle']}] {p['titel']}",
            "url": p["url"], "datum": p["datum"], "id": p["schluessel"],
        })

    eintraege.sort(key=lambda e: datum_zahl(e["datum"]), reverse=True)
    eintraege = eintraege[:60]

    jetzt = datetime.now(_ZEITZONE)
    offset = jetzt.strftime("%z") or "+0000"
    build = (f"{_WOCHENTAGE[jetzt.weekday()]}, {jetzt.day:02d} "
             f"{_MONATE[jetzt.month - 1]} {jetzt.year} "
             f"{jetzt.hour:02d}:{jetzt.minute:02d}:00 {offset}")

    teile = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<rss version="2.0">',
        "<channel>",
        "<title>Einwohnerrat Neuhausen am Rheinfall</title>",
        "<link>https://neuhausen.ch/einwohnerrat</link>",
        "<description>Vorstösse, Berichte &amp; Anträge sowie Beschlüsse &amp; "
        "Protokolle des Einwohnerrats Neuhausen am Rheinfall. "
        "Automatisch aufbereitet.</description>",
        "<language>de-ch</language>",
        f"<lastBuildDate>{build}</lastBuildDate>",
    ]
    for e in eintraege:
        teile.append("<item>")
        teile.append(f"<title>{escape(e['titel'])}</title>")
        teile.append(f"<link>{escape(e['url'])}</link>")
        teile.append(f'<guid isPermaLink="false">{escape(e["id"])}</guid>')
        pub = _rfc822(e["datum"])
        if pub:
            teile.append(f"<pubDate>{pub}</pubDate>")
        teile.append("</item>")
    teile.append("</channel>")
    teile.append("</rss>")
    return "\n".join(teile) + "\n"


def parse_aktuelles(html: str) -> list:
    """
    Aktuelles-Seite: Meldungen als Ueberschriften-Folge
      h2 = Titel, optional h3 = Untertitel, h4 = Datum (tt.mm.jjjj),
      danach Absaetze mit dem Meldungstext.
    Es wird ein kurzer Anriss (max. ~260 Zeichen) erzeugt.
    """
    soup = BeautifulSoup(html, "html.parser")
    ergebnis = []
    gesehen = set()
    aktuell = None

    def abschliessen():
        nonlocal aktuell
        if aktuell and aktuell["datum"] and aktuell["titel"]:
            anriss = re.sub(r"\s+", " ", " ".join(aktuell["absaetze"])).strip()
            if len(anriss) > 260:
                anriss = anriss[:260].rsplit(" ", 1)[0] + " \u2026"
            schluessel = f"ak|{aktuell['datum']}|{aktuell['titel']}"
            if schluessel not in gesehen:
                gesehen.add(schluessel)
                ergebnis.append({
                    "schluessel": schluessel,
                    "datum": aktuell["datum"],
                    "titel": aktuell["titel"],
                    "untertitel": aktuell["untertitel"],
                    "anriss": anriss,
                })
        aktuell = None

    for el in soup.find_all(["h2", "h3", "h4", "p"]):
        text = sauber(el.get_text())
        if el.name == "h2":
            abschliessen()
            if text:
                aktuell = {"titel": text, "untertitel": "", "datum": "",
                           "absaetze": []}
        elif aktuell is None:
            continue
        elif el.name == "h3":
            if not aktuell["datum"] and not aktuell["untertitel"]:
                aktuell["untertitel"] = text
        elif el.name == "h4":
            if DATUM_RE.match(text):
                aktuell["datum"] = text
        elif el.name == "p":
            if aktuell["datum"] and text and len(" ".join(aktuell["absaetze"])) < 400:
                aktuell["absaetze"].append(text)

    abschliessen()
    return ergebnis


# ---------------------------------------------------------------------------
# Presse: Artikel ueber Neuhausen aus regionalen Medien.
# Quellen: SHN-Rubrik Neuhausen (HTML), Schaffhauser AZ (RSS),
# Google News (RSS, Auffangnetz fuer weitere Medien).
# Es werden nur Titel, Datum, Anriss und Link uebernommen; gelesen wird
# beim jeweiligen Medium.
# ---------------------------------------------------------------------------

def _titel_schluessel(titel: str) -> str:
    """Normalisierter Titel fuer den Dubletten-Abgleich zwischen Quellen."""
    t = unicodedata.normalize("NFKD", titel or "").lower()
    t = re.sub(r"[^a-z0-9]+", "", t)
    return t[:60]


def _datum_aus_text(wert: str) -> str:
    """RFC-822- oder ISO-Datum zu tt.mm.jjjj (Schweizer Zeit)."""
    wert = (wert or "").strip()
    if not wert:
        return ""
    try:
        from email.utils import parsedate_to_datetime
        dt = parsedate_to_datetime(wert)
    except Exception:
        try:
            dt = datetime.fromisoformat(wert.replace("Z", "+00:00"))
        except Exception:
            return ""
    try:
        if dt.tzinfo is not None and _ZEITZONE is not None:
            dt = dt.astimezone(_ZEITZONE)
    except Exception:
        pass
    return dt.strftime("%d.%m.%Y")


def _kuerze_anriss(text: str, laenge: int = 240) -> str:
    import html as _html
    text = _html.unescape(_html.unescape(text or ""))
    text = sauber(re.sub(r"<[^>]+>", " ", text))
    if len(text) > laenge:
        text = text[:laenge].rsplit(" ", 1)[0] + " \u2026"
    return text


def _parse_feed(xml_text: str) -> list:
    """Minimaler RSS-2.0- und Atom-Parser (ohne Zusatzbibliotheken)."""
    import xml.etree.ElementTree as ET
    try:
        wurzel = ET.fromstring(xml_text.encode("utf-8")
                               if isinstance(xml_text, str) else xml_text)
    except Exception:
        return []

    def lokal(tag):
        return tag.rsplit("}", 1)[-1]

    eintraege = []
    for item in wurzel.iter():
        if lokal(item.tag) not in ("item", "entry"):
            continue
        titel, url, datum, anriss, quelle = "", "", "", "", ""
        for kind in item:
            name = lokal(kind.tag)
            text = (kind.text or "").strip()
            if name == "title":
                import html as _html
                titel = sauber(_html.unescape(_html.unescape(text)))
            elif name == "link":
                url = text or kind.get("href", "")
            elif name in ("pubDate", "published", "updated") and not datum:
                datum = _datum_aus_text(text)
            elif name in ("description", "summary") and not anriss:
                anriss = _kuerze_anriss(text)
            elif name == "source":
                quelle = sauber(text)
        if titel and url:
            eintraege.append({"titel": titel, "url": url.strip(),
                              "datum": datum, "anriss": anriss,
                              "quelle": quelle})
    return eintraege


def hole_presse() -> tuple:
    """Sammelt Presseartikel; gibt (eintraege, quellen_fehler) zurueck."""
    gesammelt = []
    gesehen = set()
    quellen_fehler = []

    def nimm(titel, url, datum, anriss, quelle):
        ts = _titel_schluessel(titel)
        if not ts or ts in gesehen or not datum:
            return
        gesehen.add(ts)
        gesammelt.append({
            "schluessel": f"pr|{datum}|{ts}",
            "datum": datum,
            "titel": titel,
            "quelle": quelle,
            "url": url,
            "anriss": anriss,
        })

    # 1) SHN, Rubrik Neuhausen (Datum steckt im Artikel-Link)
    try:
        for titel, url, datum in _parse_shn_liste(hole(PRESSE_SHN_URL)):
            nimm(titel, url, datum, "", "SHN")
    except Exception as e:
        quellen_fehler.append(f"SHN: {e}")

    # 2) Schaffhauser AZ (RSS), gefiltert auf Neuhausen
    try:
        for e in _parse_feed(hole(PRESSE_SHAZ_FEED)):
            if "neuhausen" not in (e["titel"] + " " + e["anriss"]).lower():
                continue
            nimm(e["titel"], e["url"], e["datum"], e["anriss"], "Schaffhauser AZ")
    except Exception as e:
        quellen_fehler.append(f"Schaffhauser AZ: {e}")

    # 3) Google News als Auffangnetz (weitere Medien)
    try:
        for e in _parse_feed(hole(PRESSE_GOOGLE_FEED)):
            titel = e["titel"]
            quelle = e["quelle"] or "weitere Medien"
            # Google haengt die Quelle mit wechselnden Trennzeichen an den Titel an
            if e["quelle"]:
                titel = re.sub(
                    r"[\s\-\u2013\u2014|\u00b7]+" + re.escape(e["quelle"]) + r"$",
                    "", titel).strip()
            nimm(titel, e["url"], e["datum"], _kuerze_anriss(e["anriss"]), quelle)
    except Exception as e:
        quellen_fehler.append(f"Google News: {e}")

    def datum_zahl(d):
        m = re.match(r"^(\d{2})\.(\d{2})\.(\d{4})$", d)
        return int(m.group(3) + m.group(2) + m.group(1)) if m else 0

    gesammelt.sort(key=lambda e: datum_zahl(e["datum"]), reverse=True)
    return gesammelt, quellen_fehler


def _parse_shn_liste(html: str) -> list:
    """Artikel-Links der SHN-Rubrik Neuhausen: [(titel, url, tt.mm.jjjj)]."""
    soup = BeautifulSoup(html, "html.parser")
    ergebnis = []
    for a in soup.find_all("a", href=True):
        if "/region/neuhausen/" not in a["href"]:
            continue
        m = re.search(r"/(\d{4})-(\d{2})-(\d{2})/", a["href"])
        if not m:
            continue
        titel = sauber(a.get_text())
        if len(titel) < 15:
            continue  # Navigations- und Weiterlesen-Links ueberspringen
        url = a["href"]
        if url.startswith("/"):
            url = "https://www.shn.ch" + url
        ergebnis.append((titel, url, f"{m.group(3)}.{m.group(2)}.{m.group(1)}"))
    return ergebnis


def _presse_datum_zahl(d: str) -> int:
    m = re.match(r"^(\d{2})\.(\d{2})\.(\d{4})$", d or "")
    return int(m.group(3) + m.group(2) + m.group(1)) if m else 0


def _lade_presse_archiv() -> dict:
    """Bestehendes Archiv aus presse_archiv.js lesen (dient als Zwischenspeicher)."""
    leer = {"rueckfuellung": False, "eintraege": {}}
    if not PRESSE_ARCHIV_AUSGABE.exists():
        return leer
    try:
        roh = PRESSE_ARCHIV_AUSGABE.read_text(encoding="utf-8")
        obj = json.loads(roh[roh.find("{"):roh.rfind("}") + 1])
        if isinstance(obj, dict) and isinstance(obj.get("eintraege"), dict):
            obj.setdefault("rueckfuellung", False)
            return obj
    except Exception:
        pass
    return leer


def _hole_lead(url: str) -> str:
    """Oeffentliche Seitenbeschreibung (og:description) eines Artikels holen."""
    soup = BeautifulSoup(hole(url), "html.parser")
    for suche in ({"property": "og:description"}, {"name": "description"}):
        meta = soup.find("meta", attrs=suche)
        if meta and meta.get("content"):
            return _kuerze_anriss(meta["content"])
    return ""


def sammle_rueckfuellung() -> tuple:
    """Einmaliger Versuch, aeltere Artikel bis PRESSE_RUECKFUELL_AB_JAHR zu holen."""
    funde = []
    meldungen = []

    # 1) SHN-Rubrik seitenweise zurueckblaettern
    try:
        seiten = 0
        for seite in range(1, 250):
            liste = _parse_shn_liste(hole(f"{PRESSE_SHN_URL}?page={seite}"))
            if not liste:
                break
            seiten += 1
            aelter = False
            for titel, url, datum in liste:
                jahr = int(datum[-4:])
                if jahr >= PRESSE_RUECKFUELL_AB_JAHR:
                    funde.append({"titel": titel, "url": url, "datum": datum,
                                  "anriss": "", "quelle": "SHN"})
                else:
                    aelter = True
            if aelter:
                break
            time.sleep(0.2)
        meldungen.append(f"SHN-Archiv: {seiten} Seiten gelesen")
    except Exception as e:
        meldungen.append(f"SHN-Archiv abgebrochen: {e}")

    # 2) Google News in Halbjahres-Fenstern
    try:
        jahr, fenster = PRESSE_RUECKFUELL_AB_JAHR, 0
        heute = datetime.now(_ZEITZONE) if _ZEITZONE else datetime.now()
        while jahr <= heute.year:
            for start, ende in ((f"{jahr}-01-01", f"{jahr}-07-01"),
                                (f"{jahr}-07-01", f"{jahr + 1}-01-01")):
                url = (PRESSE_GOOGLE_FEED
                       + f"+after:{start}+before:{ende}")
                for e in _parse_feed(hole(url)):
                    titel = e["titel"]
                    if e["quelle"]:
                        titel = re.sub(
                            r"[\s\-\u2013\u2014|\u00b7]+" + re.escape(e["quelle"]) + r"$",
                            "", titel).strip()
                    funde.append({"titel": titel, "url": e["url"],
                                  "datum": e["datum"], "anriss": _kuerze_anriss(e["anriss"]),
                                  "quelle": e["quelle"] or "weitere Medien"})
                fenster += 1
                time.sleep(0.2)
            jahr += 1
        meldungen.append(f"Google News: {fenster} Zeitfenster abgefragt")
    except Exception as e:
        meldungen.append(f"Google News abgebrochen: {e}")

    # 3) Schaffhauser AZ ueber den Such-Feed
    try:
        az_seiten = 0
        for p in range(1, 11):
            xml = hole(f"https://www.shaz.ch/?s=Neuhausen&feed=rss2&paged={p}")
            eintraege = _parse_feed(xml)
            if not eintraege:
                break
            az_seiten += 1
            for e in eintraege:
                if "neuhausen" not in (e["titel"] + " " + e["anriss"]).lower():
                    continue
                funde.append({"titel": e["titel"], "url": e["url"],
                              "datum": e["datum"], "anriss": e["anriss"],
                              "quelle": "Schaffhauser AZ"})
            time.sleep(0.2)
        meldungen.append(f"Schaffhauser AZ: {az_seiten} Suchseiten gelesen")
    except Exception as e:
        meldungen.append(f"Schaffhauser AZ abgebrochen: {e}")

    return funde, meldungen


def baue_presse() -> tuple:
    """
    Fuehrt alles zusammen: aktuelle Quellen, einmalige Rueckfuellung,
    dauerhaftes Archiv (presse_archiv.js) und Lead-Texte.
    Gibt (juengste Eintraege fuer die Seite, fehlerliste) zurueck.
    """
    archiv = _lade_presse_archiv()
    eintraege = archiv["eintraege"]
    titel_index = {}
    for s, e in eintraege.items():
        titel_index[_titel_schluessel(e.get("titel", ""))] = s

    def einpflegen(kandidat):
        ts = _titel_schluessel(kandidat["titel"])
        if not ts or not kandidat.get("datum"):
            return 0
        if ts in titel_index:
            vorhanden = eintraege[titel_index[ts]]
            if not vorhanden.get("anriss") and kandidat.get("anriss"):
                vorhanden["anriss"] = kandidat["anriss"]
            return 0
        s = f"pr|{kandidat['datum']}|{ts}"
        eintraege[s] = {
            "schluessel": s, "datum": kandidat["datum"],
            "titel": kandidat["titel"], "quelle": kandidat["quelle"],
            "url": kandidat["url"], "anriss": kandidat.get("anriss", ""),
        }
        titel_index[ts] = s
        return 1

    aktuelle, quellen_fehler = hole_presse()
    neu_aktuell = sum(einpflegen(e) for e in aktuelle)

    if not archiv.get("rueckfuellung"):
        print("  Presse: einmalige Rueckfuellung bis "
              f"{PRESSE_RUECKFUELL_AB_JAHR} laeuft (dauert etwas laenger) ...")
        rueck, meldungen = sammle_rueckfuellung()
        neu_rueck = sum(einpflegen(e) for e in rueck)
        for m in meldungen:
            print(f"    {m}")
        print(f"    Rueckfuellung: {neu_rueck} zusaetzliche Artikel erfasst")
        archiv["rueckfuellung"] = True

    # Lead-Texte: fehlende Anrisse einmalig von den Artikelseiten holen
    geholt = 0
    for e in eintraege.values():
        if e.get("anriss") or e.get("lv"):
            continue
        if "news.google.com" in e.get("url", ""):
            e["lv"] = 1  # Google-Umleitungen liefern keine Beschreibung
            continue
        if geholt >= PRESSE_LEAD_MAX_PRO_LAUF:
            continue
        try:
            e["anriss"] = _hole_lead(e["url"])
        except Exception:
            e["anriss"] = ""
        e["lv"] = 1
        geholt += 1
        time.sleep(0.15)

    js = ("window.NEUHAUSEN_PRESSE_ARCHIV = "
          + json.dumps(archiv, ensure_ascii=False) + ";\n")
    PRESSE_ARCHIV_AUSGABE.write_text(js, encoding="utf-8")

    heute = datetime.now(_ZEITZONE) if _ZEITZONE else datetime.now()
    from datetime import timedelta
    grenze = int((heute - timedelta(days=PRESSE_ANZEIGE_TAGE)).strftime("%Y%m%d"))
    alle = sorted(eintraege.values(),
                  key=lambda e: _presse_datum_zahl(e["datum"]), reverse=True)
    juengste = [e for e in alle if _presse_datum_zahl(e["datum"]) >= grenze]
    print(f"  Presse: {len(juengste)} Eintraege auf der Seite, "
          f"{len(alle) - len(juengste)} im Archiv "
          f"(neu: {neu_aktuell}, Leads geholt: {geholt})")
    return juengste, quellen_fehler


# ---------------------------------------------------------------------------
# Kennzahlen: amtliche Zahlen zur Gemeinde (Etappe 1).
# Quellen: BFS STAT-TAB (Bevoelkerung, Beschaeftigung) via oeffentliche
# Schnittstelle ohne Anmeldung, plus verifizierte Steuerfuesse des Kantons.
# Ergebnis: kennzahlen.js, hoechstens einmal pro Woche neu abgefragt.
# ---------------------------------------------------------------------------

_STATTAB_META = {}
_STATTAB_BASIS_OK = {}


def _stattab_reihe(cube: str, festlegungen: list,
                   region: str = "neuhausen am rheinfall",
                   fest: dict = None, summiere: list = None) -> tuple:
    """
    Zeitreihe fuer Neuhausen am Rheinfall aus einem STAT-TAB-Datenwuerfel.
    Liest zuerst die Struktur des Wuerfels und waehlt dann:
      - die Gemeinde Neuhausen am Rheinfall,
      - alle Jahre,
      - pro uebriger Dimension den Wert gemaess `fest` (dict: code->begriffe),
        sonst `festlegungen` (Liste von Suchbegriffen), sonst ein Total.
    Lieber gar keine Zahl als eine falsche: Ohne eindeutige Auswahl
    wird abgebrochen.
    Gibt ([(jahr, wert), ...], quellen_url) zurueck.
    """
    fest = fest or {}
    summiere = [s.lower() for s in (summiere or [])]
    basis = f"{STATTAB_BASIS}/{cube}/{cube}.px"
    if cube not in _STATTAB_META:
        r = requests.get(basis, headers=HEADERS, timeout=60)
        if r.status_code >= 400:
            # Manche Wuerfel brauchen andere Adressformen: kurz (einstufig)
            # oder dreistufig (Unterordner gleichen Namens).
            for alt_basis in (f"{STATTAB_BASIS}/{cube}.px",
                              f"{STATTAB_BASIS}/{cube}/{cube}/{cube}.px"):
                r2 = requests.get(alt_basis, headers=HEADERS, timeout=60)
                if r2.status_code < 400:
                    basis, r = alt_basis, r2
                    break
        r.raise_for_status()
        _STATTAB_META[cube] = r.json()
        _STATTAB_BASIS_OK[cube] = basis
    meta = _STATTAB_META[cube]
    basis = _STATTAB_BASIS_OK.get(cube, basis)

    def wahl(texte, begriffe):
        klein = [t.lower() for t in texte]
        for b in begriffe:
            b = b.lower()
            for i, t in enumerate(klein):
                if t == b:
                    return i
            for i, t in enumerate(klein):
                if t.startswith(b):
                    return i
            for i, t in enumerate(klein):
                if b in t:
                    return i
        return None

    def waehle_total(code, texte, werte):
        """Fuer Nicht-Regions-Dimensionen: eindeutig das Total nehmen.
        Bricht ab, wenn kein klarer Total-Wert existiert (kein Raten)."""
        i = wahl(texte, ["schweiz und ausland", "total", "gesamt",
                         "alle ", "insgesamt", "- total"])
        if i is not None:
            return werte[i]
        if len(werte) == 1:
            return werte[0]
        return None

    abfrage = []
    zeit_code = None
    for var in meta.get("variables", []):
        code = var["code"]
        texte = var.get("valueTexts", [])
        if var.get("time") or code.lower() in ("jahr", "periode"):
            zeit_code = code
            abfrage.append({"code": code,
                            "selection": {"filter": "all", "values": ["*"]}})
            continue
        ist_regions_dim = any(w in code.lower() for w in
                              ("gemeinde", "kanton", "bezirk", "region"))
        i = wahl(texte, [region]) if ist_regions_dim else None
        if i is not None:
            abfrage.append({"code": code,
                            "selection": {"filter": "item",
                                          "values": [var["values"][i]]}})
            continue
        # gezielte Festlegung einzelner Dimensionen (code -> Suchbegriffe)
        if code in fest:
            i = wahl(texte, fest[code])
            if i is None:
                raise RuntimeError(f"{cube}: Festlegung '{fest[code]}' "
                                   f"nicht gefunden in '{code}'")
            abfrage.append({"code": code,
                            "selection": {"filter": "item",
                                          "values": [var["values"][i]]}})
            continue
        # Dimension ueber alle Werte summieren (z. B. Gebaeudetyp ohne Total)
        if code.lower() in summiere:
            abfrage.append({"code": code,
                            "selection": {"filter": "all", "values": ["*"]}})
            continue
        i = wahl(texte, festlegungen)
        if i is not None:
            abfrage.append({"code": code,
                            "selection": {"filter": "item",
                                          "values": [var["values"][i]]}})
            continue
        total = waehle_total(code, texte, var["values"])
        if total is None:
            raise RuntimeError(f"{cube}: kein eindeutiges Total fuer '{code}'")
        abfrage.append({"code": code,
                        "selection": {"filter": "item", "values": [total]}})
    if zeit_code is None:
        raise RuntimeError(f"{cube}: keine Zeitachse gefunden")

    antwort = requests.post(
        basis, json={"query": abfrage, "response": {"format": "json-stat2"}},
        headers=HEADERS, timeout=90)
    antwort.raise_for_status()
    js = antwort.json()
    if "dataset" in js:  # aeltere json-stat-Variante
        js = js["dataset"]
        ids = js["dimension"]["id"]
        dims = js["dimension"]
    else:
        ids = js["id"]
        dims = js["dimension"]

    zeit_id = None
    for i in ids:
        if i.lower() == zeit_code.lower():
            zeit_id = i
            break
    if zeit_id is None:
        zeit_id = ids[-1]

    # Groessen aller Dimensionen in Reihenfolge (json-stat ist zeilen-major)
    groessen = []
    for dim_id in ids:
        kat = dims[dim_id]["category"]
        idx = kat.get("index")
        if isinstance(idx, list):
            groessen.append(len(idx))
        else:
            groessen.append(len(idx))
    werte = js["value"]

    kat = dims[zeit_id]["category"]
    index = kat.get("index")
    labels = kat.get("label", {})
    if isinstance(index, list):
        index = {c: p for p, c in enumerate(index)}
    zeit_pos_von_id = ids.index(zeit_id)

    # Fuer jede Zelle die Zeit-Koordinate bestimmen und Werte je Jahr summieren.
    def koordinate(flach):
        koords = []
        rest = flach
        for g in reversed(groessen):
            koords.append(rest % g)
            rest //= g
        return list(reversed(koords))

    jahr_von_pos = {p: str(labels.get(c, c))[:4]
                    for c, p in index.items()}
    summe = {}
    gesehen = set()
    for flach, w in enumerate(werte):
        zpos = koordinate(flach)[zeit_pos_von_id]
        jahr = jahr_von_pos.get(zpos)
        if not (jahr and re.match(r"^\d{4}$", jahr)):
            continue
        if isinstance(w, (int, float)):
            summe[jahr] = summe.get(jahr, 0) + w
            gesehen.add(jahr)
        # Strings wie "..." (Daten nicht verfuegbar) oder None -> ignorieren

    # Nur Jahre mit mindestens einem echten Wert behalten
    summe = {j: v for j, v in summe.items() if j in gesehen}

    reihe = sorted(summe.items(), key=lambda kv: int(kv[0]))
    if not reihe:
        raise RuntimeError(f"{cube}: leere Zeitreihe")
    return reihe, f"{STATTAB_SEITE}/{cube}/-/{cube}.px/"


def _wachstum(reihe: list, ab_jahr: int):
    """Prozentuale Veraenderung vom ersten Wert ab `ab_jahr` bis zum letzten."""
    ab = [(j, w) for j, w in reihe if int(j) >= ab_jahr and w]
    if len(ab) < 2 or not ab[0][1]:
        return None
    return round((ab[-1][1] / ab[0][1] - 1) * 100, 1)


def _leerwohnungsziffer(region_begriff: str = "neuhausen am rheinfall") -> tuple:
    """Holt die Leerwohnungsziffer-Zeitreihe fuer eine Gemeinde. Kaskade:
    1) BFS STAT-TAB-Wuerfel (historisch oft HTTP 400, aber falls repariert
       der direkteste Weg), 2) offene Daten via opendata.swiss-Katalog
    (Leerwohnungszaehlung als Datei). Gibt ([(jahr, ziffer), ...], url)."""
    try:
        return _leerwohnung_stattab(region_begriff)
    except Exception as e1:
        try:
            return _leerwohnung_opendata(region_begriff)
        except Exception as e2:
            raise RuntimeError(f"STAT-TAB: {e1} / opendata.swiss: {e2}")


def _leerwohnung_stattab(region_begriff: str) -> tuple:
    """Direktabruf aus dem BFS-Wuerfel px-x-0902020300_101."""
    cube = "px-x-0902020300_101"
    seite = f"{STATTAB_SEITE}/{cube}/{cube}.px"
    # Metadaten laden, mehrere URL-Formen probieren
    meta = None
    basis_ok = None
    # Dieser Wuerfel liegt in einem Unterordner gleichen Namens; seine API-
    # Adresse ist deshalb DREISTUFIG (Cube-Name dreimal). Die kuerzeren
    # Formen liefern HTTP 400, was uns lange in die Irre fuehrte.
    for basis in (f"{STATTAB_BASIS}/{cube}/{cube}/{cube}.px",
                  f"{STATTAB_BASIS}/{cube}/{cube}.px",
                  f"{STATTAB_BASIS}/{cube}.px"):
        try:
            r = requests.get(basis, headers=HEADERS, timeout=60)
            if r.status_code < 400:
                meta = r.json()
                basis_ok = basis
                break
        except Exception:
            continue
    if not meta:
        raise RuntimeError("Wuerfel nicht erreichbar")

    # Dimensionen bestimmen
    reg_var = zeit_var = None
    dim_fest = {}   # code -> gewaehlter Wert
    for var in meta["variables"]:
        code = var["code"]
        texte = var.get("valueTexts", [])
        werte = var.get("values", [])
        cl = code.lower()
        if "gemeinde" in cl or "region" in cl or "kanton" in cl:
            reg_var = var
        elif var.get("time"):
            zeit_var = var
        elif "wohnräume" in cl or "wohnraeume" in cl:
            # Total waehlen
            dim_fest[code] = _waehle_total(werte, texte)
        elif "leerwohnung" in cl and "typ" in cl:
            dim_fest[code] = _waehle_total(werte, texte)
        elif "anzahl" in cl and "anteil" in cl:
            # die Ziffer (Prozent) waehlen, nicht die absolute Anzahl
            dim_fest[code] = _waehle_begriff(
                werte, texte, ["leerwohnungsziffer", "ziffer", "anteil"])
        else:
            dim_fest[code] = _waehle_total(werte, texte)

    # Region Neuhausen finden
    reg_code = None
    for w, t in zip(reg_var["values"], reg_var["valueTexts"]):
        if region_begriff in t.lower():
            reg_code = w
            break
    if reg_code is None:
        raise RuntimeError("Region nicht gefunden")

    # Alle Jahre abfragen
    q = [{"code": reg_var["code"], "selection": {"filter": "item", "values": [reg_code]}}]
    for code, wert in dim_fest.items():
        q.append({"code": code, "selection": {"filter": "item", "values": [wert]}})
    q.append({"code": zeit_var["code"],
              "selection": {"filter": "item", "values": zeit_var["values"]}})

    r = requests.post(basis_ok, json={"query": q,
                      "response": {"format": "json-stat2"}},
                      headers=HEADERS, timeout=90)
    r.raise_for_status()
    js = r.json()
    werte = js.get("value", [])
    # Jahr-Dimension: Position -> Jahr-Label
    zeit_dim = js["dimension"][zeit_var["code"]]["category"]
    idx = zeit_dim["index"]          # code -> position
    labels = zeit_dim.get("label", {})
    pos_zu_jahr = {}
    for code, pos in idx.items():
        jahr = str(labels.get(code, code))[:4]
        if re.match(r"^\d{4}$", jahr):
            pos_zu_jahr[pos] = jahr

    reihe = []
    for pos in range(len(werte)):
        w = werte[pos]
        # Nur echte Zahlen (Platzhalter '...' kommt als None oder String)
        if isinstance(w, (int, float)) and pos in pos_zu_jahr:
            reihe.append((pos_zu_jahr[pos], round(float(w), 2)))
    reihe.sort()
    return reihe, seite


def _waehle_total(werte, texte):
    """Waehlt den Total-Wert einer Dimension (per Label 'total' oder 'alle')."""
    for w, t in zip(werte, texte):
        tl = t.lower()
        if "total" in tl or "- alle" in tl or "alle" == tl.strip():
            return w
    return werte[0] if werte else None


def _waehle_begriff(werte, texte, begriffe):
    """Waehlt den Wert, dessen Label einen der Begriffe enthaelt."""
    for w, t in zip(werte, texte):
        tl = t.lower()
        for b in begriffe:
            if b in tl:
                return w
    return werte[0] if werte else None


LEERWOHNUNG_CKAN = ("https://opendata.swiss/api/3/action/package_show"
                    "?id=leerwohnungsziffer-nach-gemeinde")


def _leerwohnung_langformat(zeilen, region_begriff):
    """Parst das lange Format: eine Zeile pro Gemeinde und Jahr, mit
    Spaltenkopf. Bestimmt Jahr- und Wertspalte aus den Koepfen; die
    Wertspalte muss eindeutig sein (Kopf mit 'ziffer'/'quote'/'anteil'),
    sonst Abbruch, damit nie Anzahlen mit Ziffern verwechselt werden."""
    kopf_idx = jahr_sp = None
    for i, row in enumerate(zeilen[:10]):
        for j, zelle in enumerate(row):
            t = str(zelle or "").lower()
            if t in ("jahr", "year", "periode", "period") or "jahr" == t[:4]:
                kopf_idx, jahr_sp = i, j
                break
        if kopf_idx is not None:
            break
    if kopf_idx is None:
        raise RuntimeError("weder Breit- noch Langformat erkannt")
    kopf = [str(z or "").lower() for z in zeilen[kopf_idx]]
    wert_sp = None
    for j, t in enumerate(kopf):
        if any(b in t for b in ("ziffer", "quote", "anteil")):
            wert_sp = j
            break
    if wert_sp is None:
        raise RuntimeError("Langformat: keine eindeutige Ziffern-Spalte")

    reihe = []
    for row in zeilen[kopf_idx + 1:]:
        rtext = " ".join(str(z) for z in row if z is not None).lower()
        if region_begriff not in rtext and not re.search(r"\b2937\b", rtext):
            continue
        if jahr_sp >= len(row) or wert_sp >= len(row):
            continue
        jj = str(row[jahr_sp]).strip()[:4]
        if not re.match(r"^(19|20)\d{2}$", jj):
            continue
        w = row[wert_sp]
        if isinstance(w, str):
            w = w.strip().replace("%", "").replace(",", ".")
            try:
                w = float(w)
            except ValueError:
                continue
        if isinstance(w, (int, float)) and 0 <= w <= 30:
            reihe.append((jj, round(float(w), 2)))
    if len(reihe) < 3:
        raise RuntimeError(f"Langformat: zu wenige plausible Werte ({len(reihe)})")
    # pro Jahr nur ein Wert (letzter gewinnt), dann sortieren
    eindeutig = {}
    for j, w in reihe:
        eindeutig[j] = w
    return (sorted(eindeutig.items()),
            "https://opendata.swiss/de/dataset/leerwohnungsziffer-nach-gemeinde")


def _leerwohnung_opendata(region_begriff: str) -> tuple:
    """Rueckfallebene: laedt die Leerwohnungsziffer aus den offenen Daten
    der Leerwohnungszaehlung (opendata.swiss-Katalog -> BFS-Datei).
    Parst die Tabelle generisch: sucht die Kopfzeile mit Jahreszahlen und
    die Gemeindezeile. Selbstvalidierend: ohne eindeutigen Fund oder bei
    unplausiblen Werten (Ziffer ausserhalb 0-30 %) wird abgebrochen,
    damit nie falsche Zahlen entstehen."""
    # 1) Katalogeintrag holen -> aktuelle Datei-URL
    r = requests.get(LEERWOHNUNG_CKAN, headers=HEADERS, timeout=60)
    r.raise_for_status()
    paket = r.json()
    if not paket.get("success"):
        raise RuntimeError("Katalogantwort ohne success")
    ressourcen = paket["result"].get("resources", [])
    # Bevorzugt Excel/CSV; erste brauchbare Ressource nehmen
    datei_url = None
    datei_format = ""
    for res in ressourcen:
        fmt = (res.get("format") or "").lower()
        url = res.get("download_url") or res.get("url") or ""
        if url and fmt in ("xlsx", "xls", "csv"):
            datei_url, datei_format = url, fmt
            break
    if not datei_url:
        raise RuntimeError("keine ladbare Ressource im Katalog")

    r = requests.get(datei_url, headers=HEADERS, timeout=120)
    r.raise_for_status()

    # 2) Tabellenzeilen gewinnen (Excel via openpyxl, CSV direkt)
    zeilen = []
    if datei_format in ("xlsx", "xls"):
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(r.content),
                                    read_only=True, data_only=True)
        # Blatt mit "ziffer" im Namen bevorzugen, sonst erstes
        blatt = None
        for name in wb.sheetnames:
            if "ziffer" in name.lower():
                blatt = wb[name]
                break
        blatt = blatt or wb[wb.sheetnames[0]]
        for row in blatt.iter_rows(values_only=True):
            zeilen.append(list(row))
    else:
        import csv
        import io
        text = r.content.decode("utf-8-sig", "replace")
        trenner = ";" if text.count(";") > text.count(",") else ","
        for row in csv.reader(io.StringIO(text), delimiter=trenner):
            zeilen.append(row)

    # 3) Kopfzeile mit Jahreszahlen finden
    def _jahr(z):
        s = str(z).strip()[:4]
        return s if re.match(r"^(19|20)\d{2}$", s) else None

    kopf_idx = None
    jahr_spalten = {}   # spaltenindex -> jahr
    for i, row in enumerate(zeilen[:30]):
        gefunden = {}
        for j, zelle in enumerate(row):
            jj = _jahr(zelle)
            if jj:
                gefunden[j] = jj
        if len(gefunden) >= 5:     # eine echte Jahres-Kopfzeile
            kopf_idx, jahr_spalten = i, gefunden
            break
    if kopf_idx is None:
        # Kein Breitformat -> Versuch im langen Format (eine Zeile pro
        # Gemeinde und Jahr), ebenfalls selbstvalidierend.
        return _leerwohnung_langformat(zeilen, region_begriff)

    # 4) Gemeindezeile finden: Namens-Treffer hat Vorrang; die BFS-Nummer
    # dient nur als Rueckfallebene (eine Zahl 2937 koennte theoretisch
    # auch als Messwert einer anderen Zeile auftauchen).
    ziel = None
    nummern_treffer = None
    for row in zeilen[kopf_idx + 1:]:
        rtext = " ".join(str(z) for z in row[:4] if z is not None).lower()
        if region_begriff in rtext:
            ziel = row
            break
        if nummern_treffer is None and re.search(r"\b2937\b", rtext):
            nummern_treffer = row
    if ziel is None:
        ziel = nummern_treffer
    if ziel is None:
        raise RuntimeError("Gemeindezeile nicht gefunden")

    # 5) Werte zuordnen und plausibilisieren
    reihe = []
    for j, jahr in sorted(jahr_spalten.items()):
        if j >= len(ziel):
            continue
        w = ziel[j]
        if isinstance(w, str):
            w = w.strip().replace("%", "").replace(",", ".")
            try:
                w = float(w)
            except ValueError:
                continue
        if isinstance(w, (int, float)) and 0 <= w <= 30:
            reihe.append((jahr, round(float(w), 2)))
    if len(reihe) < 3:
        raise RuntimeError(f"zu wenige plausible Werte ({len(reihe)})")
    reihe.sort()
    return reihe, "https://opendata.swiss/de/dataset/leerwohnungsziffer-nach-gemeinde"


def baue_kennzahlen() -> None:
    # Zwischenspeicher: hoechstens einmal pro Woche neu abfragen
    if KENNZAHLEN_AUSGABE.exists():
        try:
            roh = KENNZAHLEN_AUSGABE.read_text(encoding="utf-8")
            alt = json.loads(roh[roh.find("{"):roh.rfind("}") + 1])
            heute = int(datetime.now(_ZEITZONE).strftime("%Y%m%d")
                        if _ZEITZONE else datetime.now().strftime("%Y%m%d"))
            if (alt.get("naechste_pruefung", 0) > heute
                    and alt.get("bereiche")
                    and alt.get("version", 1) == KENNZAHLEN_VERSION):
                print("  Kennzahlen: Zwischenspeicher aktuell "
                      f"(naechste Pruefung ab {alt['naechste_pruefung']})")
                return
        except Exception:
            pass

    bereiche = []
    fehler = []

    def karte(bereich_titel, name, einheit, reihe, quelle, quelle_url,
              hinweis="", extra=None):
        for b in bereiche:
            if b["titel"] == bereich_titel:
                ziel = b
                break
        else:
            ziel = {"titel": bereich_titel, "karten": []}
            bereiche.append(ziel)
        eintrag = {
            "name": name, "einheit": einheit,
            "reihe": [list(p) for p in reihe],
            "quelle": quelle, "quelleUrl": quelle_url, "hinweis": hinweis,
        }
        if extra:
            eintrag.update(extra)
        ziel["karten"].append(eintrag)

    # --- Bevoelkerung (BFS, demografische Bilanz je Gemeinde) ---
    # Der Wuerfel schluesselt nach Staatsangehoerigkeit (Einzellaender + Total)
    # und Geschlecht auf. Wir erzwingen ueberall das Total, sonst werden
    # Kategorien addiert und die Zahl zu hoch (frueher 11'834 statt ~10'500).
    BEV_FEST = {
        "Staatsangehörigkeit (Kategorie)": ["staatsangehörigkeit (kategorie) - total"],
        "Geschlecht": ["geschlecht - total"],
        "Demografische Komponente": ["bestand am 31. dezember"],
    }
    try:
        reihe, url = _stattab_reihe(
            "px-x-0102020000_201", [], fest=BEV_FEST)
        karte("Bevölkerung", "Ständige Wohnbevölkerung", "Personen",
              reihe, "BFS, STAT-TAB", url)
        print(f"  Kennzahlen: Bevoelkerung {reihe[0][0]}\u2013{reihe[-1][0]} "
              f"({len(reihe)} Werte, aktuell {reihe[-1][1]})")
    except Exception as e:
        fehler.append(f"Bevoelkerung: {e}")

    # --- Beschaeftigung (BFS, STATENT je Gemeinde) ---
    for name, begriff in (("Beschäftigte", "beschäftigte"),
                          ("Arbeitsstätten", "arbeitsstätten")):
        try:
            reihe, url = _stattab_reihe("px-x-0602010000_102", [begriff])
            karte("Beschäftigung", name, name, reihe, "BFS, STAT-TAB", url)
            print(f"  Kennzahlen: {name} {reihe[0][0]}\u2013{reihe[-1][0]} "
                  f"({len(reihe)} Werte)")
        except Exception as e:
            fehler.append(f"{name}: {e}")

    # --- Ausländer:innenanteil (BFS, demografische Bilanz) ---
    # Derselbe Wuerfel wie die Bevoelkerung: er kennt die Kategorie "Ausland"
    # direkt. Anteil = Ausland / Gesamt, beide "Bestand am 31. Dezember",
    # Geschlecht-Total. Ergibt die amtlichen ~46 % (Gemeinde-PDF: 45.9 %).
    try:
        gem = {"Geschlecht": ["geschlecht - total"],
               "Demografische Komponente": ["bestand am 31. dezember"]}
        sak = "Staatsangehörigkeit (Kategorie)"
        ausland, url = _stattab_reihe(
            "px-x-0102020000_201", [],
            fest={**gem, sak: ["ausland"]})
        gesamt, _ = _stattab_reihe(
            "px-x-0102020000_201", [],
            fest={**gem, sak: ["staatsangehörigkeit (kategorie) - total"]})
        tot = dict(gesamt)
        reihe = []
        for jahr, aus in ausland:
            g = tot.get(jahr)
            if g and aus is not None:
                reihe.append((jahr, round(aus / g * 100, 1)))
        if reihe:
            karte("Bevölkerung", "Ausländer:innenanteil", "%",
                  reihe, "BFS, STAT-TAB", url)
            print(f"  Kennzahlen: Auslaenderanteil {reihe[0][0]}\u2013"
                  f"{reihe[-1][0]} ({len(reihe)} Werte, aktuell {reihe[-1][1]}%)")
    except Exception as e:
        fehler.append(f"Ausländer:innenanteil: {e}")

    # --- Wohnen (BFS: Neubau) ---
    # Hinweis: Die Leerwohnungsziffer (Wuerfel _101) wird weiter unten ueber
    # einen direkten, robusten Abruf geholt.
    try:
        # Aktueller Wuerfel _107 (bis 2023). Der alte _103 endete 2012.
        reihe, url = _stattab_reihe(
            "px-x-0904030000_107", [],
            fest={"Gebäudetyp": ["gebäudetyp - alle", "gebäudetyp - total"]})
        karte("Wohnen", "Neu erstellte Wohnungen", "Wohnungen",
              reihe, "BFS, STAT-TAB", url)
        print(f"  Kennzahlen: Neubau-Wohnungen {reihe[0][0]}\u2013"
              f"{reihe[-1][0]} ({len(reihe)} Werte, aktuell {reihe[-1][1]})")
    except Exception as e:
        fehler.append(f"Neu erstellte Wohnungen: {e}")

    # --- Neubau nach Zimmerzahl (Verteilung, welche Wohnungsgroessen entstehen) ---
    # Wuerfel _105: neu erstellte Wohnungen nach Zimmerzahl, bis 2024.
    # Zeigt als Verteilung, welche Groessen zuletzt gebaut wurden, plus
    # die Summe der letzten Jahre fuer eine stabilere Aussage.
    try:
        ZIMMER = [
            ("1-Zimmer", ["1-zimmer-wohnung"]),
            ("2-Zimmer", ["2-zimmer-wohnung"]),
            ("3-Zimmer", ["3-zimmer-wohnung"]),
            ("4-Zimmer", ["4-zimmer-wohnung"]),
            ("5-Zimmer", ["5-zimmer-wohnung"]),
            ("6+ Zimmer", ["6-zimmer-wohnung oder grösser"]),
        ]
        verteilung = []
        url_neubau = ""
        for label, begriffe in ZIMMER:
            reihe, url_neubau = _stattab_reihe(
                "px-x-0904030000_105", [],
                fest={"Anzahl Zimmer": begriffe})
            # Summe der letzten 5 Jahre (stabiler als ein Einzeljahr)
            letzte = reihe[-5:] if len(reihe) >= 5 else reihe
            summe = sum(w for _, w in letzte)
            verteilung.append((label, summe))
        gesamt = sum(w for _, w in verteilung)
        if gesamt > 0:
            # Als "Reihe" fuer die Karte: Label -> Anteil in Prozent
            anteile = [(label, round(w / gesamt * 100, 1))
                       for label, w in verteilung]
            karte("Wohnen", "Neubau nach Wohnungsgrösse", "%",
                  anteile, "BFS, STAT-TAB", url_neubau,
                  hinweis="Anteil der neu erstellten Wohnungen nach Zimmerzahl, "
                          "Summe der letzten 5 Jahre",
                  extra={"typ": "verteilung"})
            spitze = max(anteile, key=lambda x: x[1])
            print(f"  Kennzahlen: Neubau nach Zimmerzahl "
                  f"(haeufigste {spitze[0]} mit {spitze[1]}%)")
    except Exception as e:
        fehler.append(f"Neubau nach Zimmerzahl: {e}")

    # --- Leerwohnungsziffer (BFS Leerwohnungszaehlung, Vollerhebung) ---
    # Robuster Direktabruf: probiert mehrere URL-Formen, verkraftet den
    # BFS-Platzhalter "..." (Daten nicht verfuegbar) und die verschachtelte
    # Jahrescodierung. Fuer kleine Gemeinden schwankt die Ziffer stark,
    # daher der Hinweis.
    try:
        reihe, url = _leerwohnungsziffer()
        if reihe:
            karte("Wohnen", "Leerwohnungsziffer", "%",
                  reihe, "BFS, Leerwohnungszählung", url,
                  hinweis="Anteil leer stehender, am Markt angebotener "
                          "Wohnungen (Stichtag 1. Juni). Bei kleinen Gemeinden "
                          "von Jahr zu Jahr stark schwankend.")
            print(f"  Kennzahlen: Leerwohnungsziffer {reihe[0][0]}\u2013"
                  f"{reihe[-1][0]} ({len(reihe)} Werte, aktuell {reihe[-1][1]}%)")
    except Exception as e:
        fehler.append(f"Leerwohnungsziffer: {e}")

    # --- Vergleich mit Kanton und Schweiz (Wachstum in Prozent) ---
    for name, cube, festl, festd, ab_jahr in (
            ("Bevölkerungswachstum seit 2015",
             "px-x-0102020000_201", [], BEV_FEST, 2015),
            ("Beschäftigungswachstum seit 2011",
             "px-x-0602010000_102", ["beschäftigte"], None, 2011)):
        try:
            werte = []
            url = ""
            for label, region in (("Neuhausen", "neuhausen am rheinfall"),
                                  ("Kanton SH", "- schaffhausen"),
                                  ("Schweiz", "schweiz")):
                try:
                    reihe, url = _stattab_reihe(cube, festl, region=region,
                                                fest=festd)
                    w = _wachstum(reihe, ab_jahr)
                    if w is not None:
                        werte.append([label, w])
                except Exception:
                    continue
            if werte and werte[0][0] == "Neuhausen" and len(werte) >= 2:
                for b in bereiche:
                    if b["titel"] == "Vergleich":
                        ziel = b
                        break
                else:
                    ziel = {"titel": "Vergleich", "karten": []}
                    bereiche.append(ziel)
                ziel["karten"].append({
                    "typ": "vergleich", "name": name, "einheit": "%",
                    "werte": werte, "reihe": [],
                    "quelle": "BFS, STAT-TAB", "quelleUrl": url,
                    "hinweis": "Veränderung in Prozent im gleichen Zeitraum",
                })
                print(f"  Kennzahlen: {name}: "
                      + ", ".join(f"{l} {w:+.1f}%" for l, w in werte))
            else:
                fehler.append(f"{name}: zu wenige Regionen lieferbar")
        except Exception as e:
            fehler.append(f"{name}: {e}")

    # --- Stimmberechtigte im Verhaeltnis zur Bevoelkerung (BFS) ---
    # Wichtig: Stimmberechtigt = Schweizer ab 18. Die Quote widerspiegelt
    # daher auch Kinder- und Auslaenderanteil; das wird auf der Karte erklaert.
    try:
        stimm, url = _stattab_reihe(
            "px-x-1702020000_101", ["wahlberechtigte", "stimmberechtigte"])
        try:
            bev_reihe, _ = _stattab_reihe(
                "px-x-0102020000_201", ["bestand am 31. dezember"])
            bev = dict(bev_reihe)
        except Exception:
            bev = {}
        quote = []
        absolut = []
        for jahr, anzahl in stimm:
            if anzahl:
                absolut.append((jahr, int(anzahl)))
            # passendes Bevoelkerungsjahr suchen (gleiches oder naechstfrueheres)
            kandidaten = [j for j in bev if int(j) <= int(jahr)]
            if anzahl and kandidaten:
                bez = max(kandidaten, key=lambda j: int(j))
                if bev[bez]:
                    quote.append((jahr, round(anzahl / bev[bez] * 100, 1)))
        if absolut:
            karte("Bevölkerung", "Stimmberechtigte", "Personen",
                  absolut, "BFS, STAT-TAB (Nationalratswahlen)", url,
                  "Schweizer:innen ab 18 Jahren")
        if quote:
            karte("Bevölkerung", "Anteil Stimmberechtigte", "%",
                  quote, "BFS, STAT-TAB", url,
                  "Anteil an der Gesamtbevölkerung. Stimmberechtigt sind nur "
                  "Schweizer:innen ab 18 Jahren; der Wert widerspiegelt daher "
                  "auch den Anteil Minderjähriger und den Ausländer:innenanteil "
                  "(in Neuhausen rund 46 %).")
            print(f"  Kennzahlen: Anteil Stimmberechtigte {quote[0][0]}\\u2013"
                  f"{quote[-1][0]} ({len(quote)} Werte)")
    except Exception as e:
        fehler.append(f"Stimmberechtigte: {e}")

    # --- Finanzen (aus den Gemeinde-Jahresrechnungen, HRM2 ab 2020) ---
    # Nutzt den Finanz-Extraktor; jede Kennzahl ist durch Kontrollsummen
    # abgesichert. Enthaelt auch die Steuerfuesse als echte Zeitreihe.
    try:
        import finanz_extraktor as fx
        fin = fx.baue_finanz_zeitreihen()
        anzahl = 0
        for sch, kz in fin["kennzahlen"].items():
            reihe = kz["reihe"]  # [[jahr, wert, beurteilung], ...]
            if not reihe:
                continue
            letzte_beurteilung = reihe[-1][2] if len(reihe[-1]) > 2 else ""
            jahr = reihe[-1][0]
            pdf_url = fin.get("quellen", {}).get(
                jahr, "https://neuhausen.ch/finanzkennzahlen")
            karte("Finanzen", kz["name"], kz["einheit"],
                  [(p[0], p[1]) for p in reihe],
                  "Gemeinde Neuhausen, Jahresrechnung " + jahr, pdf_url,
                  extra={
                      "erklaerung": kz.get("erklaerung", ""),
                      "wasBedeutet": kz.get("was_bedeutet", ""),
                      "beurteilung": letzte_beurteilung,
                      "beurteilungen": [
                          [p[0], p[2] if len(p) > 2 else ""] for p in reihe],
                      "zonen": kz.get("zonen", []),
                  })
            anzahl += 1
        if anzahl:
            print(f"  Kennzahlen: Finanzen {anzahl} Kennzahlen aus "
                  f"{len(fin['jahre'])} Jahrgaengen ({', '.join(fin['jahre'])})")
    except Exception as e:
        fehler.append(f"Finanzen: {e}")

    for f in fehler:
        print(f"    Kennzahl nicht abrufbar: {f}", file=sys.stderr)

    bfs_karten = sum(len(b["karten"]) for b in bereiche
                     if b["titel"] not in ("Steuern", "Finanzen"))
    if bfs_karten == 0 and KENNZAHLEN_AUSGABE.exists():
        print("  Kennzahlen: BFS nicht erreichbar, bestehende Datei bleibt",
              file=sys.stderr)
        return

    jetzt = datetime.now(_ZEITZONE) if _ZEITZONE else datetime.now()
    from datetime import timedelta
    obj = {
        "version": KENNZAHLEN_VERSION,
        "stand": jetzt.strftime("%d.%m.%Y %H:%M"),
        "naechste_pruefung": int((jetzt + timedelta(
            days=KENNZAHLEN_PRUEFTAKT_TAGE)).strftime("%Y%m%d")),
        "bereiche": bereiche,
    }
    KENNZAHLEN_AUSGABE.write_text(
        "window.NEUHAUSEN_KENNZAHLEN = "
        + json.dumps(obj, ensure_ascii=False) + ";\n", encoding="utf-8")
    print(f"  Kennzahlen: {sum(len(b['karten']) for b in bereiche)} Karten "
          f"in {len(bereiche)} Bereichen geschrieben")


def diagnose_stattab():
    """Schreibt die echten Dimensionen und Werttexte der Kennzahlen-Wuerfel
    ins Protokoll. Aufruf: python neuhausen_daten.py --diagnose-stattab"""
    wuerfel = [
        ("Bevoelkerung", "px-x-0102020000_201"),
        ("Auslaenderanteil", "px-x-0102010000_104"),
        ("Leerwohnungen", "px-x-0902020300_101"),
        ("Neubau (aktuell)", "px-x-0904030000_107"),
    ]
    for name, cube in wuerfel:
        print(f"\n===== {name}: {cube} =====")
        try:
            basis = f"{STATTAB_BASIS}/{cube}/{cube}.px"
            r = requests.get(basis, headers=HEADERS, timeout=60)
            if r.status_code >= 400:
                r = requests.get(f"{STATTAB_BASIS}/{cube}.px",
                                 headers=HEADERS, timeout=60)
            r.raise_for_status()
            meta = r.json()
            for var in meta.get("variables", []):
                code = var["code"]
                texte = var.get("valueTexts", [])
                werte = var.get("values", [])
                ist_zeit = var.get("time") or code.lower() in ("jahr", "periode")
                if ist_zeit or len(texte) > 25:
                    # Zeit und Riesenlisten (Laender, Gemeinden) nur gekuerzt
                    proben = list(zip(werte[:4], texte[:4]))
                    print(f"  [{code}] ({len(texte)} Werte, gekuerzt): {proben} ...")
                else:
                    print(f"  [{code}] ({len(texte)} Werte):")
                    for w, t in zip(werte, texte):
                        print(f"       {w!r} = {t!r}")
        except Exception as e:
            print(f"  FEHLER: {e}")
    print("\n===== Ende Struktur-Diagnose =====")

    # --- Sonde 1: echte Bevoelkerungsabfrage (warum 11834 statt 11848?) ---
    print("\n===== SONDE Bevoelkerung 2024 =====")
    try:
        basis = f"{STATTAB_BASIS}/px-x-0102020000_201/px-x-0102020000_201.px"
        meta = requests.get(basis, headers=HEADERS, timeout=60).json()
        # Neuhausen-Code finden
        reg_code = None
        for var in meta["variables"]:
            if "gemeinde" in var["code"].lower():
                for w, t in zip(var["values"], var["valueTexts"]):
                    if "neuhausen am rheinfall" in t.lower():
                        reg_code = (var["code"], w)
            if var["code"] == "Staatsangehörigkeit (Kategorie)":
                sak_var = var
            if var["code"] == "Geschlecht":
                ge_var = var
            if var["code"] == "Demografische Komponente":
                dk_var = var
        # Total/Ausland/Schweiz-Codes und Bestand-31.12-Code ausgeben
        def zeig(var, label):
            print(f"  {label}:")
            for w, t in zip(var["values"], var["valueTexts"]):
                print(f"       code {w!r} = {t!r}")
        zeig(sak_var, "Staatsangehörigkeit (Kategorie)")
        print(f"  Neuhausen-Region: {reg_code}")
        # Abfrage bauen: Total, Geschlecht-Total, Bestand 31.12, Jahr 2024
        def code_fuer(var, begriff):
            for w, t in zip(var["values"], var["valueTexts"]):
                if begriff in t.lower():
                    return w
            return None
        q = [
            {"code": reg_code[0], "selection": {"filter": "item", "values": [reg_code[1]]}},
            {"code": "Staatsangehörigkeit (Kategorie)", "selection": {"filter": "item",
             "values": [code_fuer(sak_var, "total")]}},
            {"code": "Geschlecht", "selection": {"filter": "item",
             "values": [code_fuer(ge_var, "total")]}},
            {"code": "Demografische Komponente", "selection": {"filter": "item",
             "values": [code_fuer(dk_var, "bestand am 31. dezember")]}},
        ]
        # Jahr 2024
        for var in meta["variables"]:
            if var.get("time"):
                q.append({"code": var["code"], "selection": {"filter": "item",
                          "values": [var["values"][-1]]}})
                print(f"  Jahr (letztes): {var['values'][-1]} = {var['valueTexts'][-1]}")
        ant = requests.post(basis, json={"query": q,
              "response": {"format": "json-stat2"}}, headers=HEADERS, timeout=60)
        print(f"  Antwort-Status: {ant.status_code}")
        js = ant.json()
        print(f"  value-Array: {js.get('value')}")
        print(f"  -> erwartet 11848 laut Gemeinde-PDF")
    except Exception as e:
        print(f"  SONDE-FEHLER: {e}")

    # --- Sonde 2: Leerwohnungs-Wuerfel ueber mehrere URL-Formen ---
    print("\n===== SONDE Leerwohnungen URL-Formen =====")
    for form in (f"{STATTAB_BASIS}/px-x-0902020300_101/px-x-0902020300_101.px",
                 f"{STATTAB_BASIS}/px-x-0902020300_101.px",
                 f"{STATTAB_BASIS}/px-x-0902020300_103/px-x-0902020300_103.px"):
        try:
            r = requests.get(form, headers=HEADERS, timeout=40)
            print(f"  {form}\n     -> Status {r.status_code}")
            if r.status_code < 400:
                m = r.json()
                for var in m.get("variables", [])[:6]:
                    print(f"       [{var['code']}] {len(var.get('valueTexts', []))} Werte")
        except Exception as e:
            print(f"  {form}\n     -> FEHLER {e}")
    print("\n===== Ende Sonden =====")


def vergleichsgemeinden(min_ew=9500, max_ew=18000):
    """Ermittelt alle Schweizer Gemeinden mit einer staendigen Wohnbevoelkerung
    im angegebenen Bereich (Standard 9500-18000), aus demselben BFS-Wuerfel
    wie die Neuhausen-Bevoelkerung. Gibt eine sortierte, amtliche Liste aus.
    Aufruf: python neuhausen_daten.py --vergleichsgemeinden"""
    print(f"===== Vergleichsgemeinden {min_ew}-{max_ew} Einwohner =====")
    try:
        basis = f"{STATTAB_BASIS}/px-x-0102020000_201/px-x-0102020000_201.px"
        meta = requests.get(basis, headers=HEADERS, timeout=60).json()
    except Exception as e:
        print(f"FEHLER beim Laden der Wuerfel-Struktur: {e}")
        return

    # Dimensionen bestimmen
    reg_var = sak_var = ge_var = dk_var = zeit_var = None
    for var in meta["variables"]:
        code = var["code"]
        if "gemeinde" in code.lower():
            reg_var = var
        elif code == "Staatsangehörigkeit (Kategorie)":
            sak_var = var
        elif code == "Geschlecht":
            ge_var = var
        elif code == "Demografische Komponente":
            dk_var = var
        if var.get("time"):
            zeit_var = var

    def code_fuer(var, begriff):
        for w, t in zip(var["values"], var["valueTexts"]):
            if begriff in t.lower():
                return w
        return None

    tot_sak = code_fuer(sak_var, "total")
    tot_ge = code_fuer(ge_var, "total")
    bestand = code_fuer(dk_var, "bestand am 31. dezember")
    jahr_code = zeit_var["values"][-1]
    jahr_text = zeit_var["valueTexts"][-1]

    # Nur echte Gemeinden (Code beginnt mit "......" im valueText) auswaehlen.
    # Die Wuerfel-Struktur nutzt Praefixe: "......XXXX Gemeindename".
    gemeinde_codes = []
    gemeinde_namen = {}
    for w, t in zip(reg_var["values"], reg_var["valueTexts"]):
        if t.strip().startswith("......"):
            name = t.strip().lstrip(".").strip()
            # Fuehrende BFS-Nummer abtrennen (z. B. "2937 Neuhausen...")
            m = re.match(r"^(\d+)\s+(.*)$", name)
            if m:
                gemeinde_namen[w] = (m.group(1), m.group(2))
            else:
                gemeinde_namen[w] = ("", name)
            gemeinde_codes.append(w)

    print(f"Jahr: {jahr_text}, Gemeinden gesamt: {len(gemeinde_codes)}")
    print(f"Frage Bevoelkerung ab (das dauert einen Moment)...\n")

    # In Bloecken abfragen (alle Gemeinden auf einmal via filter "item").
    ergebnisse = []
    BLOCK = 300
    for i in range(0, len(gemeinde_codes), BLOCK):
        teil = gemeinde_codes[i:i + BLOCK]
        q = [
            {"code": reg_var["code"], "selection": {"filter": "item", "values": teil}},
            {"code": "Staatsangehörigkeit (Kategorie)",
             "selection": {"filter": "item", "values": [tot_sak]}},
            {"code": "Geschlecht", "selection": {"filter": "item", "values": [tot_ge]}},
            {"code": "Demografische Komponente",
             "selection": {"filter": "item", "values": [bestand]}},
            {"code": zeit_var["code"], "selection": {"filter": "item", "values": [jahr_code]}},
        ]
        try:
            r = requests.post(basis, json={"query": q,
                              "response": {"format": "json-stat2"}},
                              headers=HEADERS, timeout=120)
            r.raise_for_status()
            js = r.json()
            werte = js.get("value", [])
            # Reihenfolge der Region-Dimension entspricht der Abfrage-Reihenfolge
            dim = js["dimension"][reg_var["code"]]["category"]["index"]
            # index: code -> position
            pos_zu_code = {p: c for c, p in dim.items()}
            for pos in range(len(werte)):
                w = werte[pos]
                if w is None:
                    continue
                code = pos_zu_code.get(pos)
                if code in gemeinde_namen:
                    nr, name = gemeinde_namen[code]
                    if min_ew <= w <= max_ew:
                        ergebnisse.append((w, name, nr))
        except Exception as e:
            print(f"  Block {i}-{i+len(teil)}: FEHLER {e}")

    ergebnisse.sort(reverse=True)
    print(f"===== {len(ergebnisse)} Gemeinden im Bereich "
          f"{min_ew}-{max_ew} (Stand {jahr_text}) =====\n")
    print(f"{'Einwohner':>10}  {'BFS-Nr':>7}  Gemeinde")
    print("-" * 50)
    for ew, name, nr in ergebnisse:
        marker = "  <-- Neuhausen" if "neuhausen am rheinfall" in name.lower() else ""
        print(f"{ew:>10}  {nr:>7}  {name}{marker}")
    print(f"\n===== Ende Liste ({len(ergebnisse)} Gemeinden) =====")


def diagnose_wohnen_umwelt():
    """Prueft Kandidaten-Wuerfel fuer Wohnen/Umwelt auf drei Dinge:
    1) Ist der Wuerfel erreichbar? 2) Wie ist seine Struktur (Dimensionen)?
    3) Sind Neuhausen UND Kanton SH UND Schweiz als Vergleichswerte abrufbar?
    Aufruf: python neuhausen_daten.py --diagnose-wohnen"""
    # Kandidaten-Wuerfel (aus BFS-Recherche). Nummern koennen sich aendern;
    # die Diagnose zeigt bei Fehler, welche nicht erreichbar sind.
    kandidaten = [
        ("Wohnungsbestand nach Zimmerzahl", "px-x-0902020200_102"),
        ("Bewohnte Wohnungen (Bewohnertyp Miete/Eigentum)", "px-x-0903020000_123"),
        ("Neubau nach Zimmerzahl", "px-x-0904030000_105"),
        ("Arealstatistik Bodennutzung (Kandidat)", "px-x-0202020000_101"),
    ]
    for name, cube in kandidaten:
        print(f"\n{'=' * 60}")
        print(f"{name}: {cube}")
        print("=" * 60)
        # 1) Struktur laden (mit URL-Fallback wie bei den anderen Wuerfeln)
        meta = None
        for form in (f"{STATTAB_BASIS}/{cube}/{cube}.px",
                     f"{STATTAB_BASIS}/{cube}.px"):
            try:
                r = requests.get(form, headers=HEADERS, timeout=40)
                if r.status_code < 400:
                    meta = r.json()
                    break
            except Exception:
                continue
        if not meta:
            print("  NICHT ERREICHBAR (alle URL-Formen HTTP-Fehler)")
            continue

        # 2) Dimensionen zeigen (gekuerzt bei grossen Listen)
        reg_var = None
        zeit_var = None
        for var in meta["variables"]:
            code = var["code"]
            texte = var.get("valueTexts", [])
            ist_region = ("gemeinde" in code.lower() or "kanton" in code.lower()
                          or "region" in code.lower())
            if ist_region:
                reg_var = var
            if var.get("time"):
                zeit_var = var
            if ist_region or len(texte) > 20:
                print(f"  [{code}] ({len(texte)} Werte) Beispiele: "
                      f"{list(zip(var['values'][:3], texte[:3]))}")
            else:
                print(f"  [{code}] ({len(texte)} Werte):")
                for w, t in zip(var["values"], texte):
                    print(f"       {w!r} = {t!r}")

        # 3) Vergleichswerte: Neuhausen, Kanton SH, Schweiz suchen
        if reg_var:
            print("  --- Vergleichsregionen im Wuerfel vorhanden? ---")
            def suche(begriff):
                for w, t in zip(reg_var["values"], reg_var["valueTexts"]):
                    if begriff in t.lower():
                        return (w, t.strip())
                return None
            for label, begriff in (("Neuhausen", "neuhausen am rheinfall"),
                                    ("Kanton Schaffhausen", "schaffhausen"),
                                    ("Schweiz", "schweiz")):
                fund = suche(begriff)
                zeichen = "ja" if fund else "NEIN"
                extra = f" -> {fund}" if fund else ""
                print(f"       {label}: {zeichen}{extra}")
        if zeit_var:
            print(f"  Neuestes Jahr: {zeit_var['valueTexts'][-1]}")
    print(f"\n{'=' * 60}\nEnde Wohnen/Umwelt-Diagnose\n{'=' * 60}")
    print("Hinweis: 'Vergleichsregionen vorhanden' zeigt, ob spaeter ein "
          "Vergleich Neuhausen/Kanton/Schweiz moeglich ist.")


def diagnose_leerwohnung():
    """Testet die Leerwohnungsziffer-Kaskade am echten System und legt bei
    Problemen die Katalogstruktur offen.
    Aufruf: python neuhausen_daten.py --diagnose-leerwohnung"""
    print("===== Leerwohnungsziffer-Diagnose =====")
    print("\n--- Stufe 1: STAT-TAB-Wuerfel ---")
    try:
        reihe, url = _leerwohnung_stattab("neuhausen am rheinfall")
        print(f"  ERFOLG: {len(reihe)} Werte, "
              f"{reihe[0][0]}-{reihe[-1][0]}, aktuell {reihe[-1][1]}%")
    except Exception as e:
        print(f"  fehlgeschlagen: {e}")

    print("\n--- Stufe 2: opendata.swiss ---")
    try:
        r = requests.get(LEERWOHNUNG_CKAN, headers=HEADERS, timeout=60)
        print(f"  Katalog-Status: {r.status_code}")
        paket = r.json()
        ressourcen = paket.get("result", {}).get("resources", [])
        for res in ressourcen:
            print(f"    Ressource: format={res.get('format')!r} "
                  f"url={(res.get('download_url') or res.get('url') or '')[:90]}")
        # Erste ladbare Datei holen und ihren Kopf offenlegen
        for res in ressourcen:
            fmt = (res.get("format") or "").lower()
            url = res.get("download_url") or res.get("url") or ""
            if not url or fmt not in ("xlsx", "xls", "csv"):
                continue
            print(f"\n  --- Dateikopf ({fmt}) ---")
            d = requests.get(url, headers=HEADERS, timeout=120)
            print(f"  Datei-Status: {d.status_code}, {len(d.content):,} Bytes")
            if fmt in ("xlsx", "xls"):
                import io
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(d.content),
                                            read_only=True, data_only=True)
                print(f"  Blaetter: {wb.sheetnames}")
                for name in wb.sheetnames[:3]:
                    print(f"  --- Blatt {name!r}, erste 12 Zeilen ---")
                    for i, row in enumerate(wb[name].iter_rows(values_only=True)):
                        if i >= 12:
                            break
                        print(f"    {i}: {list(row)[:8]}")
            else:
                text = d.content.decode("utf-8-sig", "replace")
                for i, zeile in enumerate(text.splitlines()[:12]):
                    print(f"    {i}: {zeile[:160]}")
            break
    except Exception as e:
        print(f"  Katalog-FEHLER: {e}")
    try:
        reihe, url = _leerwohnung_opendata("neuhausen am rheinfall")
        print(f"  ERFOLG: {len(reihe)} Werte, "
              f"{reihe[0][0]}-{reihe[-1][0]}, aktuell {reihe[-1][1]}%")
        print(f"  Reihe: {reihe}")
    except Exception as e:
        print(f"  Abruf/Parsen fehlgeschlagen: {e}")
    print("\n===== Ende Leerwohnungs-Diagnose =====")


def main():
    if "--diagnose-leerwohnung" in sys.argv:
        diagnose_leerwohnung()
        return

    if "--diagnose-stattab" in sys.argv:
        diagnose_stattab()
        return

    if "--diagnose-wohnen" in sys.argv:
        diagnose_wohnen_umwelt()
        return

    if "--vergleichsgemeinden" in sys.argv:
        vergleichsgemeinden()
        return

    vorstoesse = []
    berichte = []
    beschluesse = []
    aktuelles = []
    presse = []
    fehler = []

    for quelle in VORSTOSS_QUELLEN:
        try:
            html = hole(quelle["url"])
            teil = parse_vorstoesse(html, quelle)
            vorstoesse.extend(teil)
            print(f"  {quelle['art']:16} {len(teil):>4} Eintraege")
        except Exception as e:
            fehler.append(f"{quelle['art']}: {e}")
            print(f"  {quelle['art']:16} FEHLER: {e}", file=sys.stderr)

    try:
        html = hole(BA_URL)
        berichte = parse_berichte(html)
        print(f"  Berichte&Antraege {len(berichte):>3} Eintraege")
    except Exception as e:
        fehler.append(f"Berichte & Antraege: {e}")
        print(f"  Berichte&Antraege FEHLER: {e}", file=sys.stderr)

    try:
        html = hole(BP_URL)
        beschluesse = parse_beschluesse(html)
        print(f"  Beschl.&Protokolle {len(beschluesse):>2} Eintraege")
    except Exception as e:
        fehler.append(f"Beschluesse & Protokolle: {e}")
        print(f"  Beschl.&Protokolle FEHLER: {e}", file=sys.stderr)

    try:
        html = hole(AKTUELLES_URL)
        aktuelles = parse_aktuelles(html)
        print(f"  Aktuelles         {len(aktuelles):>3} Eintraege")
    except Exception as e:
        fehler.append(f"Aktuelles: {e}")
        print(f"  Aktuelles         FEHLER: {e}", file=sys.stderr)

    try:
        presse, presse_fehler = baue_presse()
        for pf in presse_fehler:
            print(f"    Presse-Quelle nicht erreichbar: {pf}", file=sys.stderr)
        if presse_fehler and not presse:
            fehler.append("Presse: keine Quelle erreichbar")
    except Exception as e:
        print(f"  Presse FEHLER: {e}", file=sys.stderr)

    if "--ohne-kennzahlen" not in sys.argv:
        try:
            baue_kennzahlen()
        except Exception as e:
            print(f"  Kennzahlen FEHLER: {e}", file=sys.stderr)
    else:
        print("Kennzahlen uebersprungen (--ohne-kennzahlen)")

    daten = {
        "erzeugt": datetime.now(_ZEITZONE).strftime("%d.%m.%Y %H:%M"),
        "fehler": fehler,
        "vorstoesse": vorstoesse,
        "berichte": berichte,
        "beschluesse": beschluesse,
        "aktuelles": aktuelles,
        "presse": presse,
    }

    if UNBEKANNTE_PERSONEN:
        print("\nNicht zugeordnete Personen (bitte in FRAKTIONEN ergaenzen):")
        for name in sorted(UNBEKANNTE_PERSONEN):
            print(f"    \"{name}\": \"?\",")
        print()

    js = "window.NEUHAUSEN_DATEN = " + json.dumps(daten, ensure_ascii=False) + ";\n"
    AUSGABE.write_text(js, encoding="utf-8")

    feed = baue_feed(vorstoesse, berichte, beschluesse, aktuelles, presse)
    FEED_AUSGABE.write_text(feed, encoding="utf-8")

    if "--ohne-volltext" not in sys.argv:
        try:
            baue_volltext(vorstoesse, berichte, beschluesse)
        except Exception as e:
            print(f"Volltext-Erfassung fehlgeschlagen: {e}", file=sys.stderr)
    else:
        print("Volltext uebersprungen (--ohne-volltext)")

    print(f"Geschrieben: {AUSGABE}  "
          f"({len(vorstoesse)} Vorstoesse, {len(berichte)} Berichte, "
          f"{len(beschluesse)} Beschluesse, {len(aktuelles)} Aktuelles, "
          f"{len(presse)} Presse)")
    print(f"Geschrieben: {FEED_AUSGABE}")

    if fehler and not vorstoesse and not berichte:
        sys.exit(1)


if __name__ == "__main__":
    main()
